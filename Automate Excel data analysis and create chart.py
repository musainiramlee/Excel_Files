import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):  # Define a new function
    wb = xl.load_workbook(filename)  # How to load an excel file in Python
    sheet = wb['Sheet1']  # Need to specify the desired sheet

    for row in range (2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)  # We want to access column no number 3 for all rows
        corrected_price = cell.value * 0.9  # The operation that we want to do
        corrected_price_cell = sheet.cell(row, 4)  # The way to put in the result in a new column
        corrected_price_cell.value = corrected_price

    # Generate Chart and specify the desired values by selecting the column and row
    values = Reference(sheet,
              min_col=2,
              max_col=sheet.max_row,
              min_row=4,
              max_row=4)

    chart = BarChart()  # Creating Bar graph
    chart.add_data(values)  # Input in the desired 'values' for the chart
    sheet.add_chart(chart, 'f2')  # Adding the chart in excel sheet, and specify the coordinate of chart location

    wb.save(filename)  # Saving the excel file

process_workbook('transactions.xlsx')
