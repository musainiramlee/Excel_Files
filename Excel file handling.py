import openpyxl as xl
wb = xl.load_workbook('transactions.xlsx')  # How to load an excel file in Python
sheet = wb['Sheet1']  # Need to specify the desired sheet

for row in range (2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)  # We want to access column no number 3 for all rows
    corrected_price = cell.value * 0.9  # The operation that we want to do
    corrected_price_cell = sheet.cell(row, 4)  # The way to put in the result in a new column
    corrected_price_cell.value = corrected_price

wb.save('transactions2.xlsx')  # Saving the excel file